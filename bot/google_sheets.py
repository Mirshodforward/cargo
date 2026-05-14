"""Google Sheets o‘qish va Excel (.xlsx) ga eksport."""

from __future__ import annotations

import json
import logging
import math
import re
import unicodedata
from io import BytesIO
from typing import Any

from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl import Workbook

from bot.phone_utils import normalize_phone_digits

log = logging.getLogger(__name__)

SCOPES = ("https://www.googleapis.com/auth/spreadsheets",)


def _credentials(
    credentials_path: str | None,
    service_account_info: dict[str, Any] | None,
) -> Credentials:
    if service_account_info is not None:
        return Credentials.from_service_account_info(service_account_info, scopes=SCOPES)
    if credentials_path:
        return Credentials.from_service_account_file(credentials_path, scopes=SCOPES)
    raise ValueError("Google service account: .env yoki JSON fayl yo‘q")


def _service(
    credentials_path: str | None,
    service_account_info: dict[str, Any] | None,
):
    creds = _credentials(credentials_path, service_account_info)
    return build("sheets", "v4", credentials=creds, cache_discovery=False)


def fetch_values(
    spreadsheet_id: str,
    range_a1: str,
    *,
    credentials_path: str | None = None,
    service_account_info: dict[str, Any] | None = None,
) -> list[list[Any]]:
    """Jadvaldan qiymatlarni o‘qish (sinxron). asyncio.to_thread orqali chaqiring."""
    svc = _service(credentials_path, service_account_info)
    result = (
        svc.spreadsheets()
        .values()
        .get(spreadsheetId=spreadsheet_id, range=range_a1)
        .execute()
    )
    return result.get("values", [])


def _normalize_header_label(cell: Any) -> str:
    return str(cell).strip().upper()


def _cell_as_kod_string(cell_raw: Any) -> str:
    """Jadval katakchasini KOD bilan solishtirish uchun qatorga aylantiradi."""
    if cell_raw is None:
        return ""
    if isinstance(cell_raw, bool):
        return str(cell_raw)
    if isinstance(cell_raw, float) and cell_raw.is_integer():
        return str(int(cell_raw))
    if isinstance(cell_raw, int):
        return str(cell_raw)
    return str(cell_raw).strip()


def find_kod_column_index(header_row: list[Any], kod_header: str) -> int | None:
    """Birinchi qatordagi ustun nomi bo‘yicha indeks (KOD va hokazo)."""
    target = kod_header.strip().upper()
    for i, cell in enumerate(header_row):
        if _normalize_header_label(cell) == target:
            return i
    return None


def find_column_index_prefer_exact(
    header_rows: list[list[Any]],
    label: str,
) -> int | None:
    """Avvalo aniq sarlavha (NUMBER, KOD); bo‘lmasa ``find_column_index_in_header_block``."""
    for hr in header_rows:
        if not hr:
            continue
        idx = find_kod_column_index(hr, label)
        if idx is not None:
            return idx
    return find_column_index_in_header_block(header_rows, label)


def _sheet_cell_to_local_phone_digits(cell_raw: Any) -> str | None:
    """
    Jadvaldagi raqam (matn, float, yashirin belgilar, '941339383 matn) → 9 raqam.
    """
    s = _cell_as_kod_string(cell_raw)
    if not s:
        return None
    s = unicodedata.normalize("NFKC", s)
    for ch in (
        "\ufeff",
        "\u200b",
        "\u200e",
        "\u200f",
        "\u202a",
        "\u202c",
        "\xa0",
        "\u202f",
    ):
        s = s.replace(ch, "")
    s = s.strip().strip("'\"")
    if not s:
        return None
    # "941339383.0", ilmiy notatsiya (Google/Excel matn sifatida)
    if re.fullmatch(r"-?\d+(\.\d+)?([eE][+-]?\d+)?", s.replace(",", ".")):
        try:
            f = float(s.replace(",", "."))
            if math.isfinite(f) and f.is_integer() and abs(f) < 1e15:
                s = str(int(f))
        except (ValueError, OverflowError):
            pass
    return normalize_phone_digits(s)


def _pad_row_to_length(row: list[Any], min_len: int) -> list[Any]:
    if len(row) >= min_len:
        return row
    return list(row) + [""] * (min_len - len(row))


def lookup_unique_kods_by_number(
    rows: list[list[Any]],
    phone_digits: str,
    *,
    number_header: str = "NUMBER",
    kod_header: str = "KOD",
    header_row_count: int = 1,
) -> list[str]:
    """
    Birinchi ``header_row_count`` qator — sarlavha bloki.
    ``phone_digits`` — ``normalize_phone_digits`` (Telegram kontakt) natijasi.
    Mos kelgan qatorlardan KOD qiymatlari (takrorlarsiz, tartiblangan).
    """

    if not rows or len(rows) < 2:
        return []
    n_header = max(1, int(header_row_count))

    i_num: int | None = None
    i_kod: int | None = None
    data_start: int | None = None
    header_block: list[list[Any]] = []

    scan_limit = min(15, len(rows))
    for i in range(scan_limit):
        hr = rows[i]
        if not hr:
            continue
        a = find_column_index_prefer_exact([hr], number_header)
        b = find_kod_column_index_multi([hr], kod_header)
        if a is not None and b is not None:
            i_num, i_kod = a, b
            data_start = i + 1
            header_block = rows[: data_start]

    if i_num is None or i_kod is None or data_start is None:
        header_block = rows[:n_header]
        data_start = n_header
        i_num = find_column_index_prefer_exact(header_block, number_header)
        i_kod = find_column_index_prefer_exact(header_block, kod_header)
    if i_num is None or i_kod is None:
        log.warning("Lookup: «%s» yoki «%s» ustuni topilmadi.", number_header, kod_header)
        return []
    if len(rows) <= data_start:
        return []
    min_width = max(
        i_num + 1,
        i_kod + 1,
        max((len(r) for r in header_block), default=0),
    )
    seen: set[str] = set()
    out: list[str] = []
    for row in rows[data_start:]:
        padded = _pad_row_to_length(row, min_width)
        raw = padded[i_num] if i_num < len(padded) else ""
        key = _sheet_cell_to_local_phone_digits(raw)
        if not key or key != phone_digits:
            continue
        kod = _cell_as_kod_string(padded[i_kod]) if i_kod < len(padded) else ""
        if kod and kod not in seen:
            seen.add(kod)
            out.append(kod)
    out.sort(key=lambda x: (len(x), x))
    return out


def _normalize_for_match(s: str) -> str:
    """Sarlavha matni uchun normalizatsiya."""
    t = str(s).strip()
    t = unicodedata.normalize("NFKC", t).upper()
    t = t.replace("\ufeff", "").replace("\u200b", "")
    for sp in ("\xa0", "\u202f", "\u3000"):
        t = t.replace(sp, " ")
    # Barcha apostrofga o'xshash belgilar
    for ch in (
        "ʻ",
        "ʼ",
        "`",
        "\u2018",
        "\u2019",
        "\u201a",
        "\u201b",
        "\u02bc",
        "\u02bb",
    ):
        t = t.replace(ch, "'")
    t = re.sub(r"\s+", " ", t)
    return t.strip()


def _latinize_confusable_letters(s: str) -> str:
    """Katakka sarlavhada Cyrillic va Latin aralash bo'lsa (Оg'irlik)."""
    out: list[str] = []
    for c in s:
        code = ord(c)
        if code in (0x041E, 0x043E):  # О о
            out.append("O" if code == 0x041E else "o")
        elif code in (0x0421, 0x0441):  # С с (kirill C)
            out.append("C" if code == 0x0421 else "c")
        elif code in (0x0406, 0x0456):  # І і
            out.append("I" if code == 0x0406 else "i")
        else:
            out.append(c)
    return "".join(out)


def _header_match_key(s: str) -> str:
    """Apostrof va bo'shliqlarni olib tashlab solishtirish (Og'irlik / Og'irlik / Ogʻirlik)."""
    t = _normalize_for_match(s)
    t = _latinize_confusable_letters(t)
    t = re.sub(r"['`]+", "", t)
    t = re.sub(r"\s+", "", t)
    # O'g'irlik -> OGGIRLIK -> OGIRLIK
    t = re.sub(r"OG{2,}", "OG", t, flags=re.IGNORECASE).upper()
    return t


def find_column_index_for_label(header_row: list[Any], label: str) -> int | None:
    """
    Sarlavha qatorida ustun qidirish: aniq moslik, «A / B» bo‘laklari,
    yoki sarlavha qisqa yozilgan bo‘lsa prefiks (masalan «Qabul sana» → «Qabul sanasi»).
    """
    raw_label = label.strip()
    if not raw_label:
        return None
    target = _normalize_for_match(raw_label)
    target = _latinize_confusable_letters(target)
    target_key = _header_match_key(raw_label)

    for i, cell in enumerate(header_row):
        raw = str(cell).strip()
        if not raw:
            continue
        parts = [p.strip() for p in raw.split("/")]
        for p in parts:
            pn = _normalize_for_match(p)
            pn = _latinize_confusable_letters(pn)
            if pn == target:
                return i
            if _header_match_key(p) == target_key and len(target_key) >= 4:
                return i
        full = _normalize_for_match(raw)
        full = _latinize_confusable_letters(full)
        if full == target:
            return i
        if len(target_key) >= 4 and _header_match_key(raw) == target_key:
            return i
        if len(target) >= 4 and full.startswith(target):
            return i
        if len(target) >= 4 and target in full:
            return i
    return None


def find_column_index_in_header_block(
    header_rows: list[list[Any]],
    label: str,
) -> int | None:
    """Barcha sarlavha qatorlarida qidirish (Og'irlik 2-qatorda bo'lishi mumkin)."""
    for hr in header_rows:
        if not hr:
            continue
        idx = find_column_index_for_label(hr, label)
        if idx is not None:
            return idx
    target_key = _header_match_key(label)
    if len(target_key) < 3:
        return None
    for hr in header_rows:
        for i, cell in enumerate(hr):
            raw = str(cell).strip()
            if not raw:
                continue
            if _header_match_key(raw) == target_key:
                return i
    # Qisqa / kesilgan sarlavha (Excel: "Paket so...")
    for hr in header_rows:
        for i, cell in enumerate(hr):
            raw = str(cell).strip()
            if not raw:
                continue
            fk = _header_match_key(raw)
            if len(fk) < 6:
                continue
            if len(target_key) >= len(fk) and target_key.startswith(fk):
                return i
            if len(fk) >= len(target_key) and fk.startswith(target_key):
                return i
    return None


def find_status_column_index_multi(header_rows: list[list[Any]]) -> int | None:
    """Status ustuni: Status / Статус / 状态 va hokazo."""
    for label in ("Status", "STATUS", "Статус", "СТАТУС", "状态", "狀態"):
        i = find_column_index_in_header_block(header_rows, label)
        if i is not None:
            return i
    return None


def _status_cell_matches_filter(cell_raw: Any, status_filter: str) -> bool:
    """``status_filter``: ``in_transit`` | ``arrived`` | ``picked_up`` (Забрал / olib ketilgan)."""
    s = unicodedata.normalize("NFKC", str(cell_raw or "").strip())
    s = s.replace("\xa0", " ").strip()
    sl = s.lower()
    if status_filter == "in_transit":
        return "в пути" in sl
    if status_filter == "arrived":
        return sl.startswith("прибыв")
    if status_filter == "picked_up":
        return (
            "забрал" in sl
            or "забран" in sl
            or "olib ket" in sl
            or "picked up" in sl
            or "collected" in sl
            or "已取" in s
            or "领取" in s
            or "teslim al" in sl
        )
    return True


def find_kod_column_index_multi(
    header_rows: list[list[Any]],
    kod_header: str,
) -> int | None:
    """
    KOD ustuni: bir nechta sarlavha qatori (masalan КОД, 编码, KOD) ichidan qidirish.
    """
    h = (kod_header or "KOD").strip()
    if not header_rows:
        return None
    idx = find_column_index_in_header_block(header_rows, h)
    if idx is not None:
        return idx
    idx = find_column_index_prefer_exact(header_rows, h)
    if idx is not None:
        return idx
    if h.upper() == "KOD":
        for alt in ("КОД", "код", "编码", "代码", "CODE"):
            idx = find_column_index_in_header_block(header_rows, alt)
            if idx is not None:
                return idx
    return None


def _first_data_row_with_kod_in_column(
    rows: list[list[Any]],
    col_idx: int,
    kod_value: str,
    *,
    max_scan: int = 3000,
) -> int | None:
    """``col_idx`` ustunida ``kod_value`` birinchi marta uchraydigan qator indeksi."""
    needle = kod_value.strip()
    if not needle:
        return None
    limit = min(len(rows), max(1, max_scan))
    for i in range(limit):
        padded = _pad_row_to_length(rows[i], col_idx + 1)
        if col_idx >= len(padded):
            continue
        if _cell_as_kod_string(padded[col_idx]) == needle:
            return i
    return None


def parse_sheet_number(val: Any) -> float | None:
    """Vergul bilan o‘nlik sonlar (0,09594) va oddiy sonlar."""
    if val is None or val == "":
        return None
    if isinstance(val, bool):
        return None
    if isinstance(val, float) and math.isnan(val):
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("\xa0", "").replace(" ", "")
    if s in ("", "-", "—"):
        return None
    if "," in s and "." not in s:
        s = s.replace(",", ".")
    elif "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def append_umumiy_row(
    rows: list[list[Any]],
    *,
    header_row_count: int,
    sum_column_labels: tuple[str, ...],
    totals_text: str,
    totals_text_column_label: str,
) -> list[list[Any]]:
    """
    Oxiriga bitta qator: ``totals_text_column_label`` ustunida ``totals_text``,
    ``sum_column_labels`` ustunlarida faqat tanlangan KOD qatorlari yig‘indisi.
    """
    if not rows or not sum_column_labels:
        return [list(r) for r in rows]

    width = max(len(r) for r in rows)

    def pad(r: list[Any]) -> list[Any]:
        rr = list(r)
        if len(rr) < width:
            rr.extend([""] * (width - len(rr)))
        return rr

    out = [pad(list(r)) for r in rows]
    header_block = out[0:header_row_count]

    label_col = find_column_index_in_header_block(header_block, totals_text_column_label)
    if label_col is None:
        log.warning(
            "«Umumiy» matni uchun «%s» ustuni topilmadi, 2-ustun (indeks 1) ishlatiladi.",
            totals_text_column_label,
        )
        label_col = 1 if width > 1 else 0

    sum_indices: list[int] = []
    seen_ci: set[int] = set()
    for lbl in sum_column_labels:
        ci = find_column_index_in_header_block(header_block, lbl)
        if ci is None:
            log.warning("Jamlanma uchun «%s» ustuni topilmadi.", lbl)
        elif ci not in seen_ci:
            seen_ci.add(ci)
            sum_indices.append(ci)

    if not sum_indices:
        return out

    totals = [""] * width
    totals[label_col] = totals_text

    data_rows = out[header_row_count:]
    for ci in sum_indices:
        s = 0.0
        n = 0
        for row in data_rows:
            if ci < len(row):
                v = parse_sheet_number(row[ci])
                if v is not None:
                    s += v
                    n += 1
        if n:
            totals[ci] = int(round(s)) if abs(s - round(s)) < 1e-6 else round(s, 6)
        else:
            totals[ci] = ""

    out.append(totals)
    return out


def drop_column_from_rows(rows: list[list[Any]], col_idx: int) -> list[list[Any]]:
    """Har qatordan ``col_idx`` indeksidagi ustunni olib tashlaydi (.xlsx eksport)."""
    if not rows or col_idx < 0:
        return [list(r) for r in rows]
    out: list[list[Any]] = []
    for r in rows:
        rr = list(r)
        if col_idx < len(rr):
            rr.pop(col_idx)
        out.append(rr)
    return out


def build_kod_export_with_totals(
    rows: list[list[Any]],
    kod_value: str,
    *,
    kod_header: str,
    header_row_count: int,
    sum_column_labels: tuple[str, ...],
    totals_text: str,
    totals_text_column_label: str,
    status_filter: str | None = None,
) -> tuple[list[list[Any]] | None, str | None, int]:
    table, err, eff_header = filter_rows_by_kod(
        rows,
        kod_value,
        kod_header=kod_header,
        header_row_count=header_row_count,
        status_filter=status_filter,
    )
    if err or table is None:
        return None, err, eff_header
    with_totals = append_umumiy_row(
        table,
        header_row_count=eff_header,
        sum_column_labels=sum_column_labels,
        totals_text=totals_text,
        totals_text_column_label=totals_text_column_label,
    )
    kod_idx = find_kod_column_index_multi(with_totals[:eff_header], kod_header)
    if kod_idx is not None:
        with_totals = drop_column_from_rows(with_totals, kod_idx)
    return (with_totals, None, eff_header)


def filter_rows_by_kod(
    rows: list[list[Any]],
    kod_value: str,
    *,
    kod_header: str = "KOD",
    header_row_count: int = 1,
    status_filter: str | None = None,
) -> tuple[list[list[Any]] | None, str | None, int]:
    """
    Birinchi ``header_row_count`` qatori — sarlavha (filtr qo‘llanmaydi), lekin
    «KOD» ustuni pastki sarlavha qatorida bo‘lsa (КОД / 编码 / KOD) avtomatik topiladi,
    ma’lumot boshlanishi — shu KOD ustunida qiymat birinchi marta uchragan qator.

    ``status_filter``: ``in_transit`` (В пути), ``arrived`` (Прибыв…) yoki ``picked_up`` (Забрал).

    Qaytadi: ``(jadval, xato_yoki_None, sarlavha_qatorlari_soni)``.
    """
    n_param = max(1, int(header_row_count))
    if not rows:
        return None, "Jadval bo‘sh.", n_param

    needle = kod_value.strip()
    if not needle:
        return None, "KOD bo‘sh. Raqam yoki kodni yuboring.", n_param

    scan = min(len(rows), max(n_param, 20))
    header_scan = rows[:scan]
    idx = find_kod_column_index_multi(header_scan, kod_header)
    if idx is None:
        return None, (
            f"Jadval sarlavhalarida «{kod_header}» ustuni topilmadi (KOD / КОД / 编码). "
            "GOOGLE_SHEETS_HEADER_ROWS ni tekshiring (ko‘p tilli sarlavhada odatda 3)."
        ), n_param

    status_idx: int | None = None
    if status_filter:
        status_idx = find_status_column_index_multi(header_scan)
        if status_idx is None:
            return None, (
                "Jadval sarlavhalarida «Status» ustuni topilmadi (Status / Статус / 状态)."
            ), n_param

    first_hit = _first_data_row_with_kod_in_column(rows, idx, needle)
    if first_hit is not None:
        effective_header = first_hit if first_hit >= 1 else max(1, n_param)
        if first_hit < n_param:
            log.warning(
                "KOD qiymati %s-qatordan boshlanadi, GOOGLE_SHEETS_HEADER_ROWS=%s — "
                "PDF uchun sarlavha soni avtomatik moslashtirildi.",
                first_hit + 1,
                n_param,
            )
    else:
        effective_header = n_param

    if len(rows) <= effective_header:
        return None, (
            "Jadvalda ma'lumot qatorlari kam. Sarlavha qatorlari soni "
            "(GOOGLE_SHEETS_HEADER_ROWS) yoki jadval oralig‘i noto‘g‘ri bo‘lishi mumkin."
        ), effective_header

    block = rows[0:effective_header]
    data = rows[effective_header:]
    min_w = max(
        idx + 1,
        (status_idx + 1) if status_idx is not None else 0,
        max((len(r) for r in header_scan), default=0),
    )

    kod_rows: list[list[Any]] = []
    for row in data:
        padded = _pad_row_to_length(row, min_w)
        if idx >= len(padded):
            continue
        cell = _cell_as_kod_string(padded[idx])
        if cell == needle:
            kod_rows.append(row)

    if not kod_rows:
        return None, f"«{needle}» KOD bo‘yicha qator topilmadi.", effective_header

    if status_filter and status_idx is not None:
        filtered: list[list[Any]] = []
        for row in kod_rows:
            padded = _pad_row_to_length(row, min_w)
            raw_st = padded[status_idx] if status_idx < len(padded) else ""
            if _status_cell_matches_filter(raw_st, status_filter):
                filtered.append(row)
        if not filtered:
            return None, (
                "Tanlangan status bo‘yicha qatorlar yo‘q (В пути / Прибывшие / Забрал)."
            ), effective_header
        kod_rows = filtered

    matched: list[list[Any]] = list(block) + kod_rows
    return matched, None, effective_header


def _http_error_detail(exc: HttpError) -> str:
    try:
        payload = json.loads(exc.content.decode("utf-8"))
        return str(payload.get("error", {}).get("message", ""))
    except Exception:
        return str(exc)


def user_message_for_sheets_error(exc: Exception) -> str:
    """Telegram foydalanuvchisiga tushunarli xabar (keng tarqalgan API xatolari)."""
    if isinstance(exc, HttpError):
        status = exc.resp.status if exc.resp else 0
        detail = _http_error_detail(exc)
        log.warning("Google Sheets API: status=%s detail=%s", status, detail)

        if status == 403:
            return (
                "Jadvalga ruxsat yo‘q. Google Sheetsda «Ulashish» → JSON kalitdagi "
                "service account emailini qo‘shing (kamida «Ko‘rish» / Viewer)."
            )
        if status == 404:
            return (
                "Jadval topilmadi: GOOGLE_SPREADSHEET_ID noto‘g‘ri yoki fayl "
                "boshqa akkauntga tegishli."
            )
        if status == 400:
            return (
                "Jadval oralig‘i yoki varaq nomi noto‘g‘ri. .env da "
                "GOOGLE_SHEETS_EXPORT_RANGE ni Google Sheetsdagi varaq nomiga "
                "moslang (masalan: «Лист1!A:ZZ» yoki «Data!A1:Z»)."
            )
        log.warning("Google Sheets API: boshqa HTTP xato status=%s", status)
        return (
            "Jadvalni yuklab bo‘lmadi. Keyinroq urinib ko‘ring yoki "
            "administrator bilan bog‘laning."
        )
    log.exception("Google Sheets: kutilmagan xato")
    return (
        "Jadvalni yuklab bo‘lmadi. Keyinroq urinib ko‘ring yoki "
        "administrator bilan bog‘laning."
    )


def rows_to_xlsx_bytes(
    rows: list[list[Any]],
    *,
    worksheet_name: str = "Yuk",
) -> bytes:
    """2D qatorlardan .xlsx fayl baytlari."""
    wb = Workbook()
    ws = wb.active
    ws.title = (worksheet_name or "Sheet")[:31]

    for ri, row in enumerate(rows, start=1):
        for ci, val in enumerate(row, start=1):
            ws.cell(row=ri, column=ci, value="" if val is None else val)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def sheet_title_to_a1_range(sheet_title: str, cell_range: str = "A:ZZ") -> str:
    """Varaq nomi uchun A1 notatsiya (maxsus belgilar va apostrof)."""
    inner = str(sheet_title).replace("'", "''")
    cr = (cell_range or "A:ZZ").strip() or "A:ZZ"
    return f"'{inner}'!{cr}"


def _sanitize_excel_sheet_title(raw: str) -> str:
    t = str(raw).strip() or "Sheet"
    for ch in r':\/*?[]':
        t = t.replace(ch, "_")
    t = t[:31] or "S"
    return t


def _unique_worksheet_titles(raw_titles: list[str]) -> list[str]:
    used_lower: set[str] = set()
    out: list[str] = []
    for raw in raw_titles:
        base = _sanitize_excel_sheet_title(raw)
        name = base
        n = 1
        while name.lower() in used_lower:
            suffix = f"_{n}"
            max_base = max(1, 31 - len(suffix))
            name = (base[:max_base] + suffix)[:31]
            n += 1
        used_lower.add(name.lower())
        out.append(name)
    return out


def fetch_spreadsheet_sheet_titles(
    spreadsheet_id: str,
    *,
    credentials_path: str | None = None,
    service_account_info: dict[str, Any] | None = None,
) -> list[str]:
    """Jadvaldagi barcha varaqlar (UI tartibi bo‘yicha)."""
    svc = _service(credentials_path, service_account_info)
    meta = (
        svc.spreadsheets()
        .get(
            spreadsheetId=spreadsheet_id,
            fields="sheets.properties",
        )
        .execute()
    )
    pairs: list[tuple[int, str]] = []
    for sh in meta.get("sheets", []):
        prop = sh.get("properties") or {}
        title = str(prop.get("title") or "Sheet").strip() or "Sheet"
        try:
            idx = int(prop.get("index", 0))
        except (TypeError, ValueError):
            idx = 0
        pairs.append((idx, title))
    pairs.sort(key=lambda x: x[0])
    return [t for _, t in pairs]


def batch_get_values_ranges(
    spreadsheet_id: str,
    ranges: list[str],
    *,
    credentials_path: str | None = None,
    service_account_info: dict[str, Any] | None = None,
) -> list[list[list[Any]]]:
    """``ranges`` tartibida har bir oralik uchun qiymatlar jadvali."""
    if not ranges:
        return []
    svc = _service(credentials_path, service_account_info)
    chunk_size = 50
    grids: list[list[list[Any]]] = []
    for offset in range(0, len(ranges), chunk_size):
        chunk = ranges[offset : offset + chunk_size]
        result = (
            svc.spreadsheets()
            .values()
            .batchGet(spreadsheetId=spreadsheet_id, ranges=chunk)
            .execute()
        )
        vrs = result.get("valueRanges", [])
        if len(vrs) != len(chunk):
            log.warning(
                "batchGet: chunk %s so‘rov, valueRanges %s",
                len(chunk),
                len(vrs),
            )
        for j in range(len(chunk)):
            if j < len(vrs):
                grids.append(vrs[j].get("values") or [])
            else:
                grids.append([])
    return grids


def full_spreadsheet_to_xlsx_bytes(
    spreadsheet_id: str,
    *,
    cell_range: str = "A:ZZ",
    credentials_path: str | None = None,
    service_account_info: dict[str, Any] | None = None,
) -> bytes:
    """
    Barcha varaqlarni bitta .xlsx da: har bir varaq uchun ``cell_range`` (masalan A:ZZ).
    """
    titles = fetch_spreadsheet_sheet_titles(
        spreadsheet_id,
        credentials_path=credentials_path,
        service_account_info=service_account_info,
    )
    if not titles:
        raise ValueError("Jadvalda varaq topilmadi.")
    ranges = [sheet_title_to_a1_range(t, cell_range) for t in titles]
    grids = batch_get_values_ranges(
        spreadsheet_id,
        ranges,
        credentials_path=credentials_path,
        service_account_info=service_account_info,
    )
    ws_names = _unique_worksheet_titles(titles)

    wb = Workbook()
    first = True
    for rows, ws_title in zip(grids, ws_names):
        if first:
            ws = wb.active
            ws.title = ws_title
            first = False
        else:
            ws = wb.create_sheet(title=ws_title)
        for ri, row in enumerate(rows, start=1):
            for ci, val in enumerate(row, start=1):
                ws.cell(row=ri, column=ci, value="" if val is None else val)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
